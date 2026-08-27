"""
Convert the Roma Dashboard Data Extract (xlsx) into one JSON file per country tab.

Each sheet is a stack of "indicator blocks". A block looks like:

    Row 0: <id>                 <description>
    Row 1: 'formula'            <metaData text>
    Row 2: 'base'               <base/universe text>
    Row 3: (blank)
    Row 4: (blank col A)  'Total' 'Population' ''  'Roma - Sex...' '' 'Non-Roma - Sex...' ...
    Row 5: (blank col A)  'Total' 'Roma' 'Non-Roma' 'Female' 'Male' 'Female' 'Male' ...
    Row 6: 'n='                 <counts per column>
    Row 7+: 'Yes'/'No'/'Total'  OR  'Mean'  OR  'Gap (pp)'   <values per column>
    Row N: (blank)  -> next block starts

Column layout (fixed across every block in this workbook):
    col 2 = overall total (population, Roma+Non-Roma combined)
    col 3 = Roma, no breakdown
    col 4 = Non-Roma, no breakdown
    col 5+ = pairs of merged groups: "Roma - <disaggregation>" then
             "Non-Roma - <disaggregation>", each spanning 2-4 sub-columns
             (the categories, given on the row right below the group row).

This script walks each block, uses the two header rows to figure out which
ethnicity + disaggregation + category every column belongs to, and emits:

{
  "id": ...,
  "description": ...,
  "metaData": ...,
  "total": <overall n=>,
  "roma": [ {disaggregation, category, noOfRespondents, yesPercent, noPercent} | {..., mean} | {..., gapPp}, ... ],
  "nonRoma": [ ... ]
}
"""

import json
import re
from pathlib import Path

import openpyxl

SOURCE_XLSX = "./Roma_Dashboard_Data_Extract.xlsx"
OUTPUT_DIR = Path("./")

# Row-1 labels that are NOT indicator ids - used to recognise where a new
# indicator block starts.
RESERVED_LABELS = {"n=", "Yes", "No", "Total", "Mean", "Gap (pp)", "formula", "base", None}

# How to shorten the long "Roma - X of the respondent" style group headers
# down to a short disaggregation slug.
DISAGGREGATION_MAP = {
    "Sex": "sex",
    "Age groups": "age group",
    "Urbanisation": "urbanisation",
    "Educational attainment": "education",
    "Employment status": "employment status",
    "Household size": "household size",
    "Children in the household": "children in household",
    "Age (WB)": "age (wb)",
}

SUFFIX_RE = re.compile(r" of the (respondent|person|household head)$", re.IGNORECASE)


def normalize_disaggregation(raw_group_text: str) -> str:
    """'Roma \u2014 Sex of the respondent' -> 'sex'."""
    # Drop the leading 'Roma \u2014 ' / 'Non-Roma \u2014 ' prefix.
    label = raw_group_text.split("\u2014", 1)[-1].strip()
    label = SUFFIX_RE.sub("", label).strip()
    return DISAGGREGATION_MAP.get(label, label.lower())


def is_roma_group(raw_group_text: str) -> bool:
    text = raw_group_text.strip()
    if text.startswith("Non-Roma"):
        return False
    if text.startswith("Roma"):
        return True
    raise ValueError(f"Unrecognised group header: {raw_group_text!r}")


def to_number(value):
    """Blank cells (e.g. n=0 groups) become 0 instead of null."""
    return value if value is not None else 0


def parse_block(ws, id_row):
    """Parse one indicator block starting at row `id_row`. Returns (dict, next_row)."""
    id_ = ws.cell(row=id_row, column=1).value
    description = ws.cell(row=id_row, column=2).value
    meta_data = ws.cell(row=id_row + 1, column=2).value  # 'formula' row

    # Skip blank row(s) to find the group-header row.
    r = id_row + 3
    while all(ws.cell(row=r, column=c).value is None for c in range(1, 45)):
        r += 1
    group_row, sub_row, n_row = r, r + 1, r + 2
    assert ws.cell(row=n_row, column=1).value == "n=", (
        f"Expected 'n=' at row {n_row}, got {ws.cell(row=n_row, column=1).value!r} "
        f"(block starting row {id_row})"
    )

    # Collect every value row (Yes/No/Total, or Mean, or Gap (pp)) until a blank row.
    value_rows = []
    vr = n_row + 1
    while ws.cell(row=vr, column=1).value is not None:
        value_rows.append((ws.cell(row=vr, column=1).value, vr))
        vr += 1
    next_row = vr

    # Figure out how many columns this block actually uses (usually 44, i.e. up to AR).
    max_col = 4
    for c in range(5, 45):
        if ws.cell(row=sub_row, column=c).value is not None:
            max_col = c

    # Build per-column metadata: (is_roma, disaggregation, category)
    col_info = {}
    current_group = None
    for c in range(5, max_col + 1):
        header = ws.cell(row=group_row, column=c).value
        if header is not None:
            current_group = header
        category = ws.cell(row=sub_row, column=c).value
        col_info[c] = (is_roma_group(current_group), normalize_disaggregation(current_group), category.lower())

    # Grab a named value row by label, or None if this block doesn't have it.
    def row_values(label):
        for lbl, row_idx in value_rows:
            if lbl == label:
                return {c: ws.cell(row=row_idx, column=c).value for c in range(2, max_col + 1)}
        return None

    n_values = {c: ws.cell(row=n_row, column=c).value for c in range(2, max_col + 1)}
    yes_values = row_values("Yes")
    no_values = row_values("No")
    mean_values = row_values("Mean")
    gap_values = row_values("Gap (pp)")

    def build_entry(col):
        entry = {
            "noOfRespondents": to_number(n_values.get(col)),
        }
        if yes_values is not None:
            entry["yesPercent"] = to_number(yes_values.get(col))
            entry["noPercent"] = to_number(no_values.get(col))
        elif mean_values is not None:
            entry["mean"] = to_number(mean_values.get(col))
        elif gap_values is not None:
            entry["gapPp"] = to_number(gap_values.get(col))
        return entry

    roma_list = [{"disaggregation": "none", "category": "none", **build_entry(3)}]
    non_roma_list = [{"disaggregation": "none", "category": "none", **build_entry(4)}]

    for c in range(5, max_col + 1):
        roma_flag, disaggregation, category = col_info[c]
        entry = {"disaggregation": disaggregation, "category": category, **build_entry(c)}
        (roma_list if roma_flag else non_roma_list).append(entry)

    block = {
        "id": id_,
        "description": description,
        "metaData": meta_data,
        "noOfRespondents": to_number(n_values.get(2)),
        "roma": roma_list,
        "nonRoma": non_roma_list,
    }
    return block, next_row


def parse_sheet(ws):
    indicators = []
    row = 1
    max_row = ws.max_row
    while row <= max_row:
        val = ws.cell(row=row, column=1).value
        if val is not None and val not in RESERVED_LABELS:
            block, next_row = parse_block(ws, row)
            indicators.append(block)
            row = next_row
        else:
            row += 1
    return indicators


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        indicators = parse_sheet(ws)
        out_path = OUTPUT_DIR / f"{sheet_name.lower()}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(indicators, f, ensure_ascii=False, indent=2)
        print(f"{sheet_name}: {len(indicators)} indicators -> {out_path}")


if __name__ == "__main__":
    main()